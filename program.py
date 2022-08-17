"""
Written by Fadil Isamotu (v1.0)
November 16, 2021
moisa1@morgan.edu
"""
import pyvisa
import numpy  
import os
from time import sleep as idle 
from time import time as time_Elapsed 
import datetime                 
#import matplotlib.pyplot as plt 
import csv
import openpyxl
from openpyxl import load_workbook
import math
from shutil import copyfile
import json


m = M = min  = MIN  = minute = MINUTE = minutes = MINUTES = 60
h = H = hour = HOUR = hours  = HOURS  =  3600
d = D = day  = DAY  = days   = DAYS   =  86400

PROGRAM_START_TIME    = time_Elapsed()
dT_Diode_Continuous = 0
dT_DEVICE_1 = 0
dT_DEVICE_2 = 0
dT_DEVICE_3 =0
counter =0

run_DEVICE_1_IV = None
run_DEVICE_2_IV = None
run_DEVICE_3_IV = None
#-----------------------------------------------------------------------------------------------------Time Unit Convertions-----------------------------------------------------------------------------------------------

# Function adapted from Harshad Surdi (Arizona State University)
def time_passed(deltaT):
    hrs=round(deltaT/3600,1)
    mins=round(deltaT/60,1)
    secs=round(deltaT,1)
    return [hrs,mins,secs]

#-----------------------------------------------------------------------------------------------------Argument Parser-----------------------------------------------------------------------------------------------------
def arg(file):
    with open(f"{file}.json", "r") as arguments:
        return (json.load(arguments)) 
#-----------------------------------------------------------------------------------------------------Create log file's directory----------------------------------------------------------------------------------------- 
if not os.path.exists( log_D:= (arg("diode_IVs_Args")["directory"]) +"\log_Book"):       
    os.makedirs(log_D)
log_Book_Name ="log book.CSV"
# If workbook does not exist, create one
with open( os.path.join(log_D,  log_Book_Name) , 'w', newline='' ) as myfile:   
    writer = csv.writer(myfile)
    writer.writerow(["Date", "Time elapsed", "Measurement", "Measurement duration", "Set Frequency"])  
    myfile.close()             

#------------------------------------------------------------------------------------------------PYVISA INITIALIZATION-------------------------------------------------------------------------------------------------
# pyvisa's resource manager to get devices Id
rm = pyvisa.ResourceManager()
#print(rm.list_resources()) #__Optional__ To check what visa devices are available
mtrx = rm.open_resource('USB0::0x0957::0x3D18::MY56480010::0::INSTR') # Switching matrix ID
smu  = rm.open_resource('USB0::0x0957::0x8C18::MY51140489::0::INSTR') # SMU ID

#-----------------------------------------------------------------------------------------------------SWITCHING MATRIX----------------------------------------------------------------------------------------------------
def matrix_Open(device):
    """This function opens specified channels on DB25 Breakout 25 Pin Connector using U2751A switching matrix.  
    
    >>> matrix_Open(1)
    Opens channels for transistor 1

    >>> matrix_Open("all")
    Opens all channels

    >>> matrix_Open("RST")
    Resets the switching matrix (Opens all channels)
    """  
    
    if device == 'all' or device == 'ALL' or device == 'All':
        
        mtrx.write(f'ROUTe:OPEN (@ 101, 202, 103, 204, 105, 206, 107, 208)')
        
    elif device == "RST" or device == "RESET" or device == "reset" or device == "rst":
        
        mtrx.write('*RST')    
        
    elif device == '1':
        
        mtrx.write(f'ROUTe:OPEN (@ 101, 202)')
        
    elif device == '2':
        
        mtrx.write(f'ROUTe:OPEN (@ 103, 204)')
        
    elif device == '3':
        
        mtrx.write(f'ROUTe:OPEN (@ 105, 206)')
    
    elif device == '4':
        mtrx.write(f'ROUTe:OPEN (@ 107,208)')

    mtrx.write('opc?')                       # Waits for all commands to be executed before executing next commands   

def matrix_Close(device): 
    """This function closes specified channels on DB25 Breakout 25 Pin Connector using U2751A switching matrix.  
    
    >>> matrix_Open(1)
    Closes channels for transistor 1

    >>> matrix_Open("all")
    Closes all channels

    >>> matrix_Close("RST")
    Resets the switching matrix (Opens all channels)
    """ 
    
    if device == 'all' or device == 'ALL' or device == 'All':
        
        mtrx.write(f'ROUTe:CLOSE (@ 101, 202, 103, 204, 105, 206, 107, 208)')
        
    elif device == "RST" or device == "RESET" or device == "reset" or device == "rst":
        
        mtrx.write('*RST')    
        
    elif device == '1':
        
        mtrx.write(f'ROUTe:CLOSE (@ 101, 202)')
        
    elif device == '2':
        
        mtrx.write(f'ROUTe:CLOSE (@ 103, 204)')
        
    elif device == '3':
        
        mtrx.write(f'ROUTe:CLOSE (@ 105, 206)')
    
    elif device == '4':
        mtrx.write(f'ROUTe:CLOSE (@ 107,208)')
    
    mtrx.write('opc?')                       # Waits for all commands to be executed before executing next commands    

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------SWEEPS----------------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------SWEEPS----------------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------SWEEPS----------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------IDVD SWEEP------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def ID_VD(state = "on", message = '', VG_Start = 0, VG_End = 5, VDS_Start = 0,  VDS_END = 5,
    VG_Data_Points = 11,  VDS_Data_Points = 101, device_Name ='',
    ID_COMPLIANCE = 12e-2, ID_RESOLUTION = 0.01, IG_COMPLIANCE=1e-3, 
    IG_RESOLUTION = 0.01, directory = "D:\TESTS", folder_Name = 'IDVD_SWEEP_', device_Channel = '1',  measurement = 'ID_VD'):

    """This function performs ID/VD sweeps using channel 1 of B2902A SMU for drain bias and channel 2 for gate bias.
    A CSV file of the biased and measured data are saved in a directory specified by the "directory" parameter.
    The directory' and files' names contain the date and time during which they were created. 

    "Start" and "End" parameters respectively specify at what numeric value to start and end the sweep.

    The "datapoints" parameters specify in how many increments/slices/steps the corresponding argument must be divided.
    datapoints = [(End value)/(Increment/Step value)] + 1 (The first data point is the value of “start”.)

    Ex1: VG/VD_start = 0, VG/VD_End = 18.  
    VG/VD_Data_Points = 7,  ==> step = 3      ==>  [0.0, 3.0, 6.0, 9.0, 12.0, 15.0, 18.0]
    VG/VD_Data_Points = 10, ==> step = 2      ==>  [0.0, 2.0, 4.0, 6.0, 8.0, 10.0, 12.0, 14.0, 16.0, 18.0]

    Ex2: VG/VD_start = 4, VG/VD_End = 48.
    VG/VD_Data_Points = 5,  ==> step = 11      ==>  (4.0, 15.0, 26.0, 37.0, 48.0)
    VG/VD_Data_Points = 4,  ==> step = 14.66   ==>  (4.0, 18.666666666666664, 33.33333333333333, 48.0)
    VG/VD_Data_Points = 6,  ==> step = 8.8     ==>  (4.0, 12.8, 21.6, 30.400000000000002, 39.2, 48.0) 

    Resolution parameters (PLC) affect the measurement speed and quality.
    The highest speed setting (0.01) results in increased noise and lower measurement accuracy, whereas the lowest speed (10) results in best measurement quality.
    """
    if state != "on": return

    start_Time = time_Elapsed()
    global dT_DEVICE_1
    global dT_DEVICE_2
    global dT_DEVICE_3
    print(message)

    today = datetime.datetime.today()   
    with open( os.path.join(log_D, log_Book_Name) , 'a', newline='' ) as myfile:   
        writer = csv.writer(myfile)
        writer.writerow([f'{today:%m_%d_%Y_%H_%M}', str(time_passed((time_Elapsed() 
        - PROGRAM_START_TIME))[0]) + "H", f"Starting {measurement}"])
        myfile.close() 

    # Empty lists to store output data
    Vds         = []
    sensed_VDS  = []
    Vg          = []
    sensed_VG   = []
    Ig          = []
    Id          = []

    smu.write('*rst')                                   # Resets the sourcemeter
    smu.write('opc?')                                   # Waits for all commands to be executed before executing next commands
    mtrx.write('*RST')                                  # Resets the switching matrix
    mtrx.write('opc?')                                  # Waits for all commands to be executed before executing next commands
    matrix_Close(device_Channel)                        # Closing specified channel

    # Channel 1 configuration

    smu.write(":SENS1:CURR:RANG:AUTO ON")               # Setting channel 1's sensing range to automatic mode
    smu.write(f':sens1:func:ALL')                       # Setting channel to sense any queried value (Voltage, Current, Resistance)
    smu.write(f':sens1:curr:nplc {ID_RESOLUTION}')      # Measurement Resolution
    smu.write(f':sens1:curr:prot {ID_COMPLIANCE}')      # Drain to source current Compliance
    smu.write(":SOUR1:FUNC:MODE VOLT")                  # Setting channel 1 to source voltage
    smu.write(":OUTP ON")                               # Turning channel 1 on

    # Channel 2 configuration

    smu.write(":SENS2:CURR:RANG:AUTO ON")               # Setting channel 2's sensing range to automatic mode
    smu.write(f':sens2:func:ALL')                       # Setting channel to sense any queried value (Voltage, Current, Resistance)
    smu.write(f':sens2:curr:nplc {IG_RESOLUTION}')      # Measurement Resolution
    smu.write(f':sens2:curr:prot {IG_COMPLIANCE}')      # Gate current Compliance
    smu.write(":SOUR2:FUNC:MODE VOLT")                  # Setting channel 2 to source voltage
    smu.write(":OUTP ON")                               # Turning channel 2 on
    smu.write('opc?')                                   # Waits for all commands to be executed before executing next commands

    # Loops/Sweeps

    for VG in numpy.linspace(VG_Start, VG_End, num=VG_Data_Points):          # Outer loop alters VG
        
        #smu.write(f':SYST:BEEP 300, 0.4')              # Beep at every VG increment
        smu.write(":SOUR2:VOLT " + str(VG))             # Sources programmed value of VG from channel 2

        smu.write('opc?')                               # Waits for all commands to be executed before executing next commands

        for VDS in numpy.linspace(VDS_Start, VDS_END, num=VDS_Data_Points): # Inner loop alters VD

            Vds.append(VDS)                             # Appends programmed value of VDS to VDS's list 
            Vg.append(VG)                               # Appends programmed value of VG to VG's list      
            smu.write(":SOUR1:VOLT " + str(VDS) )       # Sources programmed value of VDS from channel 1
           
            # Getting sensed Gate, and Drain Source voltages
            smu.write(f':meas:VOLT? (@1)')                # Queries voltage value measured at channel 1
            sensed_VDS.append(smu.read().rstrip("\n"))    # Removes newline from acquired value and appends it to sensed_VDS

            smu.write(f':meas:VOLT? (@2)')                # Queries voltage value measured at channel 2
            sensed_VG.append(smu.read().rstrip("\n"))     # Removes newline from acquired value and appends it to sensed_VG
    
            # Getting sensed current 
            smu.write(f':meas:curr? (@1)')              # Queries current value measured at channel 1
            Id.append(smu.read().rstrip("\n"))          # Removes newline from acquired value and appends it to Id
        
            smu.write(f':meas:curr? (@2)')              # Queries current value measured at channel 2
            Ig.append(smu.read().rstrip("\n"))          # Removes newline from acquired value and appends it to Ig
            
            smu.write('opc?')                           # Waits for all commands to be executed before executing next commands
            
            
    #smu.write(f':SYST:BEEP 400, 1')                    # Longbeep for 1 second after sweeps are over

    # Converting all sensed values from strings to floats
    Ig          = [eval(x) for x in Ig]
    Id          = [eval(x) for x in Id]
    sensed_VDS  = [eval(x) for x in sensed_VDS]
    sensed_VG   = [eval(x) for x in sensed_VG]

    # Folder to save files
    today = datetime.datetime.today()                               # Indexing current date and time with variable "today" 
    export_Folder = (f"{directory}\{folder_Name}{today:%m-%d-%Y}")  # Formatting export folder
    if not os.path.exists(export_Folder):                           # To make sure that no additional folder is created if saving folder already exists
        os.makedirs(export_Folder)

    # Writing results to a .CSV file and saving it to created folder
    IDVD_CSV = (f"{device_Name}_IDVD_{today:%m-%d-%Y-%H_%M}.CSV")                    # Formating file's name
    with open( os.path.join( export_Folder,  IDVD_CSV) , 'w', newline='' ) as myfile: # Saving CSV file in created directory    
        writer = csv.writer(myfile)
        writer.writerow(["VDS","sensed_VDS", "ID", "IG", "VG","sensed_VG"])               # File's header
        for v in range (0, VG_Data_Points*VDS_Data_Points):                               # Setting range to read all values in lists
            writer.writerow ([ Vds[v], sensed_VDS[v], Id[v], Ig[v], Vg[v], sensed_VG[v]]) # Specifying rows
        myfile.close() 

    
    today = datetime.datetime.today()   
    with open( os.path.join(log_D, log_Book_Name) , 'a', newline='' ) as myfile:   
        writer = csv.writer(myfile)
        writer.writerow([f'{today:%m_%d_%Y_%H_%M}', str(time_passed((time_Elapsed() 
        - PROGRAM_START_TIME))[0]) + "H", f"Ending {measurement}",  str(int(time_Elapsed()- start_Time)) + "s",  arg("IV_to_Run_Next_Args")[device_Name + "_IV_Frequency"] ])
        myfile.close() 


    dT_DEVICE_3 = dT_DEVICE_3 + int (time_Elapsed()- start_Time)
    if device_Name   == "DEVICE_2":  dT_DEVICE_1 = dT_DEVICE_1 + int (time_Elapsed()- start_Time)
    elif device_Name == "DEVICE_1":  dT_DEVICE_2 = dT_DEVICE_2 + int (time_Elapsed() - start_Time)
    
    print("DEVICE_1 time passed " + str(dT_DEVICE_1))
    print("DEVICE_2 time passed " + str(dT_DEVICE_2))
    print("DEVICE_3 time passed " + str(dT_DEVICE_3))



#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------IDVG SWEEP------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def ID_VG(state = "on", message = '', VG_Start = -3,  VG_End = 5, VDS_Start = 0.5, VDS_END = 3,
    VG_Data_Points = 242, VDS_Data_Points = 6, device_Name ='',
    ID_COMPLIANCE = 12e-2, ID_RESOLUTION = 0.01, IG_COMPLIANCE = 1e-3,
    IG_RESOLUTION = 0.01, directory= "D:\TESTS", folder_Name = 'IDVG_SWEEP_', device_Channel ='1', measurement = 'ID_VG'):

    """This function performs ID/VG sweeps using channel 1 of B2902A SMU for drain bias and channel 2 for gate bias.
    The bias consists of a forward and backward gate voltage sweep (From VG_Start to VG_End, then from VG_End to VG_Start).
    A CSV file of the biased and measured data are saved in a directory specified by the "directory" parameter.
    The directory' and files' names contain the date and time during which they were created. 


    "Start" and "End" parameters respectively specify at what numeric value to start and end the sweep.


    The "datapoints" parameters specify in how many increments/slices/steps the corresponding argument must be divided.
    datapoints = [(End value)/(Increment/Step value)] + 1 (The first data point is the value of “start”.)

    Ex1: VG/VD_start = 0, VG/VD_End = 18.  
    VG/VD_Data_Points = 7,  ==> step = 3      ==>  [0.0, 3.0, 6.0, 9.0, 12.0, 15.0, 18.0]
    VG/VD_Data_Points = 10, ==> step = 2      ==>  [0.0, 2.0, 4.0, 6.0, 8.0, 10.0, 12.0, 14.0, 16.0, 18.0]

    Ex2: VG/VD_start = 4, VG/VD_End = 48.
    VG/VD_Data_Points = 5,  ==> step = 11      ==>  (4.0, 15.0, 26.0, 37.0, 48.0)
    VG/VD_Data_Points = 4,  ==> step = 14.66   ==>  (4.0, 18.666666666666664, 33.33333333333333, 48.0)
    VG/VD_Data_Points = 6,  ==> step = 8.8     ==>  (4.0, 12.8, 21.6, 30.400000000000002, 39.2, 48.0)

    Resolution parameters (PLC) affect the measurement speed and quality.
    The highest speed setting (0.01) results in increased noise and lower measurement accuracy, whereas the lowest speed (10) results in best measurement quality.
    """ 
    if state != "on": return
    start_Time = time_Elapsed()
    global dT_DEVICE_1
    global dT_DEVICE_2
    global dT_DEVICE_3
    print(message)
    HALF_OF_VG_DATAPOINTS = int(VG_Data_Points/2)

    today = datetime.datetime.today()   
    with open( os.path.join(log_D, log_Book_Name) , 'a', newline='' ) as myfile:   
        writer = csv.writer(myfile)
        writer.writerow([f'{today:%m_%d_%Y_%H_%M}', str(time_passed((time_Elapsed() 
        - PROGRAM_START_TIME))[0]) + "H", f"Starting {measurement}"])
        myfile.close() 

    # Creating empty lists to store output data
    Vds        = []
    sensed_VDS = []
    Vg         = []
    sensed_VG  = []
    Ig         = []
    Id         = []

    smu.write('*rst')                                   # Resets the sourcemeter
    smu.write('opc?')                                   # Waits for all commands to be executed before executing next commands
    mtrx.write('*RST')                                  # Resets the switching matrix
    mtrx.write('opc?')                                  # Waits for all commands to be executed before executing next commands
    matrix_Close(device_Channel)                        # Closing specified channel

    # Channel 1 config

    smu.write(":SENS1:CURR:RANG:AUTO ON")               # Setting channel 1's sensing range to automatic mode
    smu.write(f':sens1:func ""curr""')                  # Setting channel to sense current
    smu.write(f':sens1:curr:nplc {ID_RESOLUTION}')      # Measurement Resolution
    smu.write(f':sens1:curr:prot {ID_COMPLIANCE}')      # Drain to source current Compliance
    smu.write(":SOUR1:FUNC:MODE VOLT")                  # Setting channel 1 to source voltage
    smu.write(":OUTP ON")                               # Turning channel 1 on

    # Channel 2 config

    smu.write(":SENS2:CURR:RANG:AUTO ON")               # Setting channel 2's sensing range to automatic mode
    smu.write(f':sens2:func ""curr""')                  # Setting channel to sense current
    smu.write(f':sens2:curr:nplc {IG_RESOLUTION}')      # Measurement Resolution
    smu.write(f':sens2:curr:prot {IG_COMPLIANCE}')      # Gate current Compliance
    smu.write(":SOUR2:FUNC:MODE VOLT")                  # Setting channel 2 to source voltage
    smu.write(":OUTP ON")                               # Turning channel 2 on

    smu.write('opc?')                                   # Waits for all commands to be executed before executing next commands

    # Setting up loops 

    for VDS in numpy.linspace(VDS_Start, VDS_END, num = VDS_Data_Points):           # Outer loop alters VD
        
        #smu.write(f':SYST:BEEP 300, 0.4')                                           # Beep at every VD increment
        smu.write(":SOUR1:VOLT " + str(VDS))                                        # Sources programmed value of VDS

        smu.write('opc?')                                                           # Waits for all commands to be executed before executing next commands
            
        for VG_UP in numpy.linspace(VG_Start, VG_End, num = HALF_OF_VG_DATAPOINTS): # First inner loop alters VG increments
            # Appends set values of VG and VDS to their specific lists
            Vds.append(VDS)                                                         # Appends programmed value of VDS to VDS's list 
            Vg.append(VG_UP)                                                        # Appends programmed value of VG to VG's list 
            smu.write(":SOUR2:VOLT " + str(VG_UP) )                                 # Sources programmed value of VG from channel 2

            # Getting sensed Gate, and Drain Source voltages
            smu.write(f':meas:VOLT? (@1)')              # Queries voltage value measured at channel 1
            sensed_VDS.append(smu.read().rstrip("\n"))  # Removes newline from acquired value and appends it to sensed_VDS
            smu.write(f':meas:VOLT? (@2)')              # Queries voltage value measured at channel 2
            sensed_VG.append(smu.read().rstrip("\n"))   # Removes newline from acquired value and appends it to sensed_VG
            
            # Getting sensed current 
            smu.write(f':meas:curr? (@1)')              # Queries current value measured at channel 1
            Id.append(smu.read().rstrip("\n"))          # Removes newline from acquired value and appends it to Id
            smu.write(f':meas:curr? (@2)')              # Queries current value measured at channel 2
            Ig.append(smu.read().rstrip("\n"))          # Removes newline from acquired value and appends it to Ig
            
            smu.write('opc?')                           # Waits for all commands to be executed before executing next commands
        
        for VG_DOWN in numpy.linspace(VG_End, VG_Start, num = HALF_OF_VG_DATAPOINTS): # Second inner loop alters VG decrements
            Vds.append(VDS)
            Vg.append(VG_DOWN)
            
            smu.write(":SOUR2:VOLT " + str(VG_DOWN))    # Sources programmed value of VDS
            
            # Getting sensed Gate, and Drain Source voltages
            smu.write(f':meas:VOLT? (@1)')              # Queries voltage value measured at channel 1
            sensed_VDS.append(smu.read().rstrip("\n"))    # Removes newline from acquired value and appends it to sensed_VDS
            smu.write(f':meas:VOLT? (@2)')              # Queries voltage value measured at channel 2
            sensed_VG.append(smu.read().rstrip("\n"))     # Removes newline from acquired value and appends it to sensed_VG

            smu.write(f':meas:curr? (@1)')              # Queries current value measured at channel 1
            Id.append(smu.read().rstrip("\n"))          # Removes newline from acquired value and appends it to Id
            smu.write(f':meas:curr? (@2)')              # Queries current value measured at channel 2
            Ig.append(smu.read().rstrip("\n"))          # Removes newline from acquired value and appends it to Ig
           
            smu.write('opc?')                           # Waits for all commands to be executed before executing next commands
        
            
    #smu.write(f':SYST:BEEP 400, 1')                     # Longbeep for 1 second after sweeps are over

    
    # Converting all sensed values from strings to floats
    Ig          = [eval(x) for x in Ig]
    Id          = [eval(x) for x in Id]
    sensed_VDS  = [eval(x) for x in sensed_VDS]
    sensed_VG   = [eval(x) for x in sensed_VG]

    # Creating lists with absolute value of IG and ID for IDVG plot
    Ig_ABSL = [abs(num) for num in Ig]
    Id_ABSL = [abs(num) for num in Id]

    # Folder to save files
    today = datetime.datetime.today()                                                                             # Indexing current date and time with variable "today" 
    export_Folder = (f"{directory}\{folder_Name}{today:%m-%d-%Y}")                                                # Formatting export folder  
    if not os.path.exists(export_Folder):                                                                         # To make sure that no additional folder is created is saving folder already exists
        os.makedirs(export_Folder)

    # Writing results to a .CSV file and saving it to created folder
    IDVG_CSV = (f"{device_Name}_IDVG_{today:%m-%d-%Y-%H_%M}_{time_Elapsed() - start_Time}.CSV")                                                     # Formating file's name
    with open( os.path.join( export_Folder,  IDVG_CSV) , 'w', newline='' ) as myfile:                                 # Saving CSV file in created directory
        writer = csv.writer(myfile)
        writer.writerow(["VG", "sensed_VG", "ID_abs", "IG_abs", "Id", "IG", "VDS", "sensed_VDS"])                     # File's header
        for v in range (0, VG_Data_Points*VDS_Data_Points):                                                           # Setting range to read all values in lists
            writer.writerow ([ Vg[v], sensed_VG[v], Id_ABSL[v], Ig_ABSL[v], Id[v], Ig[v], Vds[v], sensed_VDS[v] ])    # Specifying rows
        myfile.close() 

    today = datetime.datetime.today()   
    with open( os.path.join(log_D, log_Book_Name) , 'a', newline='' ) as myfile:   
        writer = csv.writer(myfile)
        writer.writerow([f'{today:%m_%d_%Y_%H_%M}', str(time_passed((time_Elapsed() 
        - PROGRAM_START_TIME))[0]) + "H", f"Ending {measurement}", str(int(time_Elapsed()- start_Time)) + "s", 
         arg("IV_to_Run_Next_Args")[device_Name + "_IV_Frequency"]])
        myfile.close() 
        
    dT_DEVICE_3 = dT_DEVICE_3 + int (time_Elapsed()- start_Time)
    if device_Name   == "DEVICE_2":  dT_DEVICE_1 = dT_DEVICE_1 + int (time_Elapsed()- start_Time)#; run_DEVICE_2_IV = False
    elif device_Name == "DEVICE_1":  dT_DEVICE_2 = dT_DEVICE_2 + int (time_Elapsed()- start_Time)

    print("DEVICE_1 time passed " + str(dT_DEVICE_1))
    print("DEVICE_2 time passed " + str(dT_DEVICE_2))
    print("DEVICE_3 time passed " + str(dT_DEVICE_3))

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------IV SWEEPS------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def diode_IVs( state = "on", message = '', V_Start = 0, V_END_pt1 = -2.8, V_END_pt2 = 2.8,  V_Data_Points = 1001, 
            pt1_compliance = 50e-3, pt2_compliance = 50e-3 , rsl = 0.01,
            directory = "D:\TESTS", folder_Name = 'IVs_Diode', IV_Pt1_File_Name = "17-050_GEER800411-450C_<hrs_value>hrs_fIV.xlsx", 
            IV_Pt2_File_Name = "17-050_GEER800411-450C_<hrs_value>hrs_rIV.xlsx",
            device_Channel = '3',  measurement = 'Diode\'s IVs'):

    """This function performs IV sweeps using channel 1 of B2902A SMU.
    The bias consists of a forward and backward gate voltage sweep (From VG_Start to VG_End, then from VG_End to VG_Start).
    A CSV file of the biased and measured data are saved in a directory specified by the "directory" parameter.

    Ex1: V_start = 0, V_End = 18.  
    V_Data_Points = 7,  ==> step = 3      ==>  [0.0, 3.0, 6.0, 9.0, 12.0, 15.0, 18.0]
    V_Data_Points = 10, ==> step = 2      ==>  [0.0, 2.0, 4.0, 6.0, 8.0, 10.0, 12.0, 14.0, 16.0, 18.0]

    Ex2: V_start = 4, V_End = 48.
    V_Data_Points = 5,  ==> step = 11      ==>  (4.0, 15.0, 26.0, 37.0, 48.0)
    V_Data_Points = 4,  ==> step = 14.66   ==>  (4.0, 18.666666666666664, 33.33333333333333, 48.0)
    V_Data_Points = 6,  ==> step = 8.8     ==>  (4.0, 12.8, 21.6, 30.400000000000002, 39.2, 48.0) 
    """
    if state != "on": return
    
    start_Time = time_Elapsed()
    global dT_DEVICE_2
    global dT_DEVICE_1
    print(message)

    today = datetime.datetime.today()   
    with open( os.path.join(log_D, log_Book_Name) , 'a', newline='' ) as myfile:   
        writer = csv.writer(myfile)
        writer.writerow([f'{today:%m_%d_%Y_%H_%M}', str(time_passed((time_Elapsed() 
        - PROGRAM_START_TIME))[0]) + "H", f"Starting {measurement}"])
        myfile.close() 


    # Lists for IV part 1
    IV_pt1_voltage               = []
    sensed_IV_pt1_voltage        = []
    IV_pt1_Current               = []

    # Lists for IV part2
    IV_pt2_voltage               = []
    sensed_IV_pt2_voltage        = []
    IV_pt2_Current               = []


    smu.write('*rst')                                   # Resets the sourcemeter
    smu.write('opc?')                                   # Waits for all commands to be executed before executing next commands
    mtrx.write('*RST')                                  # Resets the switching matrix
    mtrx.write('opc?')                                  # Waits for all commands to be executed before executing next commands
    matrix_Close(device_Channel)                        # Closing specified channel

    smu.write(":SENS1:CURR:RANG:AUTO ON")                       # Setting channel 1's sensing range to automatic mode
    smu.write(f':sens1:func ""curr""')                          # Setting channel to sense current
    smu.write(f':sens1:curr:nplc {rsl}')                        # Measurement Resolution
    smu.write(f':sens1:curr:prot {pt1_compliance}')             # current Compliance
    smu.write(":SOUR1:FUNC:MODE VOLT")                          # Setting channel 1 to source IV_pt1_voltage 
    smu.write(":OUTP ON")                                       # Turning channel 1 on
    smu.write('opc?')                                           # Waits for all commands to be executed before executing next commands 


    for V in numpy.linspace(V_Start, V_END_pt1, num = V_Data_Points): 

        IV_pt1_voltage.append(V)                                   # Appends programmed value of IV_pt1_voltage to IV_pt1_voltage's list   
        smu.write(":SOUR1:VOLT " + str(V) )                        # Sources programmed value of IV_pt1_voltage from channel 1
        
        # Getting sensed voltage 
        smu.write(f':meas:VOLT? (@1)')                             # Queries voltage value measured at channel 1
        sensed_IV_pt1_voltage.append(smu.read().rstrip("\n"))      # Removes newline from acquired value and appends it to IV_pt1_voltage's list

        # Getting sensed current 
        smu.write(f':meas:curr? (@1)')                             # Queries current value measured at channel 1
        IV_pt1_Current.append(smu.read().rstrip("\n"))             # Removes newline from acquired value and appends it to IV_pt1_Current's list

        smu.write('opc?')                                          # Waits for all commands to be executed before executing next commands

    # Converting all sensed values from strings to floats
    sensed_IV_pt1_voltage  = [eval(x) for x in sensed_IV_pt1_voltage ]
    IV_pt1_Current         = [eval(x) for x in IV_pt1_Current]
    hours_Passed_Pt1       = str ( round ( time_passed(time_Elapsed() - PROGRAM_START_TIME)[0]) )


    for V in numpy.linspace(V_Start, V_END_pt2, num = V_Data_Points): 
        
        smu.write(f':sens1:curr:prot {pt2_compliance}')            # current Compliance
        IV_pt2_voltage.append(V)                                   # Appends programmed value of IV_pt2_voltage to IV_pt2_voltage's list
        smu.write(":SOUR1:VOLT " + str(V) )                        # Sources programmed value of IV_pt2_voltage from channel 1
        
        # Getting sensed IV_pt1_voltage 
        smu.write(f':meas:VOLT? (@1)')                             # Queries voltage value measured at channel 1
        sensed_IV_pt2_voltage.append(smu.read().rstrip("\n"))      # Removes newline from acquired value and appends it to IV_pt2_voltage's list

        # Getting sensed current 
        smu.write(f':meas:curr? (@1)')                             # Queries current value measured at channel 1
        IV_pt2_Current.append(smu.read().rstrip("\n"))             # Removes newline from acquired value and appends it to IV_pt2_Current's list
        smu.write('opc?')                                          # Waits for all commands to be executed before executing next commands

    # Converting all sensed values from strings to floats
    sensed_IV_pt2_voltage  = [eval(x) for x in sensed_IV_pt2_voltage ]
    IV_pt2_Current         = [eval(x) for x in IV_pt2_Current]


    # Formatting export folders  

    export_Folder = (f"{directory}\{folder_Name}")                                                  # Formatting export folder  
    if not os.path.exists(export_Folder):                                                           # To make sure that no additional folder is created is saving folder already exists
        os.makedirs(export_Folder)

    hours_Passed_Pt2 = str(round(time_passed(time_Elapsed() - PROGRAM_START_TIME)[0]))

    IV_pt1_Export_File = (f"{export_Folder}\{IV_Pt1_File_Name.replace('<hrs_value>', hours_Passed_Pt1 )}")          
    IV_pt2_Export_File = (f"{export_Folder}\{IV_Pt2_File_Name.replace('<hrs_value>', hours_Passed_Pt2 )}")      
    current_Column_Pt1 = "17-050_GEER800411-450C_<hrs_value>hrs".replace('<hrs_value>', hours_Passed_Pt1)
    voltage_Column_Pt1 = "17-050_GEER800411-450C_<hrs_value>hrs_V".replace('<hrs_value>', hours_Passed_Pt1)
    current_Column_Pt2 = "17-050_GEER800411-450C_<hrs_value>hrs".replace('<hrs_value>', hours_Passed_Pt2)
    voltage_Column_Pt2 = "17-050_GEER800411-450C_<hrs_value>hrs_V".replace('<hrs_value>', hours_Passed_Pt2)    

    if (not os.path.isfile(IV_pt1_Export_File)):        
        wb2 = openpyxl.Workbook()
        sheet = wb2.active
        c1 = sheet.cell(row = 1, column = 1)
        c2 = sheet.cell(row = 1, column = 2)

        # writing values to cells
        (c1.value, c2.value) = (
        current_Column_Pt1, 
        voltage_Column_Pt1)
        wb2.save(IV_pt1_Export_File)

    if (not os.path.isfile(IV_pt2_Export_File)):        
        wb3 = openpyxl.Workbook()
        sheet = wb3.active
        c1 = sheet.cell(row = 1, column = 1)
        c2 = sheet.cell(row = 1, column = 2)

        # writing values to cells
        (c1.value, c2.value) = (
        current_Column_Pt2, 
        voltage_Column_Pt2)
        wb3.save(IV_pt2_Export_File)

    wb2 = load_workbook(IV_pt1_Export_File)
    ws2 = wb2.worksheets[0]

    for row in range(len(IV_pt1_Current)):
        ws2.append((IV_pt1_Current[row], IV_pt1_voltage [row]))
    wb2.save(IV_pt1_Export_File)
    wb2.close()
    

    wb3 = load_workbook(IV_pt2_Export_File)
    ws3 = wb3.worksheets[0]
    
    for row in range(len(IV_pt2_Current)):
        ws3.append((IV_pt2_Current[row], IV_pt2_voltage [row]))
    wb3.save(IV_pt2_Export_File)
    wb3.close()

    today = datetime.datetime.today()   
    with open( os.path.join(log_D, log_Book_Name) , 'a', newline='' ) as myfile:   
        writer = csv.writer(myfile)
        writer.writerow([f'{today:%m_%d_%Y_%H_%M}', str(time_passed((time_Elapsed() 
        - PROGRAM_START_TIME))[0]) + "H", f"Ending {measurement}",  str(int(time_Elapsed()- start_Time)) + "s",  arg("IV_to_Run_Next_Args")["DEVICE_3_IV_Frequency"]])
        myfile.close() 
    
    dT_DEVICE_1 = dT_DEVICE_1 + int (time_Elapsed()- start_Time)
    dT_DEVICE_2 = dT_DEVICE_2 + int (time_Elapsed()- start_Time)


#----------------------------------------------------------------------------------------------------------MEDEVICE_3REMENTS TRIGGERS------------------------------------------------------------------------------------------------------

def check_What_IV_to_Run_Next (DEVICE_1_IV_Frequency, DEVICE_2_IV_Frequency, DEVICE_3_IV_Frequency, freq_Margin):

    global run_DEVICE_1_IV 
    run_DEVICE_1_IV              = None
    global run_DEVICE_2_IV
    run_DEVICE_2_IV         = None
    global run_DEVICE_3_IV
    run_DEVICE_3_IV              = None
    
    global dT_DEVICE_1

    global dT_DEVICE_2
    
    global dT_DEVICE_3
         
    
    if  ( DEVICE_1_IV_Frequency - dT_DEVICE_1  <=  freq_Margin):
        run_DEVICE_1_IV = True
        dT_DEVICE_1 = 0
      
    if  ( DEVICE_2_IV_Frequency - dT_DEVICE_2 <= freq_Margin):
        run_DEVICE_2_IV = True
        dT_DEVICE_2 = 0

    if  ( DEVICE_3_IV_Frequency - dT_DEVICE_3 <=  freq_Margin):
        run_DEVICE_3_IV = True
        dT_DEVICE_3 = 0
    
    print ("DEVICE_1 countdown = {}, state {}, time passed {}".format( DEVICE_1_IV_Frequency - dT_DEVICE_1, run_DEVICE_1_IV, dT_DEVICE_1))
    print ("DEVICE_2 countdown= {}, state {}, time passed {}".format( DEVICE_2_IV_Frequency - dT_DEVICE_2, run_DEVICE_2_IV, dT_DEVICE_2))
    print ("DEVICE_3 countdown = {}, state {}, time passed {}\n\n".format(DEVICE_3_IV_Frequency - dT_DEVICE_3, run_DEVICE_3_IV, dT_DEVICE_3))
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------CONSTANT VOLTAGE------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def diode_Continuous (state = "on", message = '', screen = "off", current_Sense_Frequency = .5, continuous_Voltage = 2.8, rsl = 0.01,
                directory = "D:\TESTS", continuous_File_Name = '17-050_GEER800411-450C_tIV.xlsx', 
                current_Compliance = 50e-3, device_channel = '3', backup_Time_in_Minutes = 5, measurement = "Diode'\s constant"):

    if state != "on": return
    print(message)
    start_Time = time_Elapsed()
    smu.write('*rst')                                   # Resets the sourcemeter
    smu.write('opc?')                                   # Waits for all commands to be executed before executing next commands
    mtrx.write('*RST')                                  # Resets the switching matrix
    mtrx.write('opc?')                                  # Waits for all commands to be executed before executing next commands
    matrix_Close(device_channel)                        # Closing specified channel

    global counter
    global dT_DEVICE_1
    global dT_DEVICE_2 
    global dT_DEVICE_3
 
    today = datetime.datetime.today()   
    with open( os.path.join(log_D, log_Book_Name) , 'a', newline='' ) as myfile:   
        writer = csv.writer(myfile)
        writer.writerow([f'{today:%m_%d_%Y_%H_%M}', str(time_passed((time_Elapsed() 
        - PROGRAM_START_TIME))[0]) + "H", f"Starting {measurement}"])
        myfile.close() 

    # Formatting export folders
    continuous_Export_File     = (f"{directory}\{continuous_File_Name}") 

    # If workbook does not exist, create one.
    if (not os.path.isfile(continuous_Export_File)):        
        wb1 = openpyxl.Workbook()
        sheet = wb1.active
        c1 = sheet.cell(row = 1, column = 1)
        c2 = sheet.cell(row = 1, column = 2)
        c3 = sheet.cell(row = 1, column = 3)
        c4 = sheet.cell(row = 1, column = 4)
        c5 = sheet.cell(row = 1, column = 5)
        c6 = sheet.cell(row = 1, column = 6)

        # writing values to cells
        (c1.value, c2.value, c3.value, c4.value, c5.value, c6.value) = (
            "hrs", "min", "sec", "volt_prgm", "volt_sense", "curr_sense")
        wb1.save(continuous_Export_File)
        counter = time_Elapsed()


    smu.write(":SENS1:CURR:RANG:AUTO ON")                       # Setting channel 1's sensing range to automatic mode
    smu.write(f':sens1:func ""curr""')                          # Setting channel to sense current
    smu.write(f':sens1:curr:nplc {rsl}')                        # Measurement Resolution
    smu.write(f':sens1:curr:prot {current_Compliance}')         # current Compliance
    smu.write(":SOUR1:FUNC:MODE VOLT")                          # Setting channel 1 to source IV_pPROGRAM_START_TIME_voltage 
    smu.write(":OUTP ON")                                       # Turning channel 1 on
    smu.write('opc?')                                           # Waits for all commands to be executed before executing next commands 
    smu.write(":SOUR1:VOLT " + str(continuous_Voltage) )        # Sources programmed value of V from channel 1

    if arg("diode_Continuous_Args")["screen"] == "off": smu.write(":DISP:ENAB OFF")
    elif arg("diode_Continuous_Args")["screen"] == "on": smu.write(":DISP:ENAB ON")

    MD, MG, SD, SG, AI = (arg("DEVICE_1_IDVD_Args")["state"], arg("DEVICE_1_IDVG_Args")["state"],
    arg("DEVICE_2_IDVD_Args")["state"], arg("DEVICE_2_IDVG_Args")["state"],arg("diode_IVs_Args")["state"])
    
    while True:
        wb1 = load_workbook(continuous_Export_File)
        ws1 = wb1.worksheets[0]

        # Getting sensed voltage 
        smu.write(f':meas:VOLT? (@1)')                               # Queries voltage value measured at channel 1
        sensed_Continuous_Voltage = eval(smu.read().rstrip("\n"))    # Removes newline from acquired value and save it in variable

        # Getting sensed current 
        smu.write(f':meas:curr? (@1)')                               # Queries current value measured at channel 1
        current = eval(smu.read().rstrip("\n"))                      # Removes newline from acquired value and save it in variable
        
        tp = time_passed(time_Elapsed() - PROGRAM_START_TIME)        # Getting time elapsed

        ws1.append( ( tp[0], tp[1], tp[2], continuous_Voltage,
        sensed_Continuous_Voltage, current ) )

        wb1.save(continuous_Export_File)
        wb1.close()

        td = datetime.datetime.today()                               # Indexing current date and time with variable "td" 
        td = f'{td:%m_%d_%Y_%H_%M}'

        if ( (tst := (time_Elapsed()-counter))/60 > backup_Time_in_Minutes):
            copyfile(continuous_Export_File, continuous_Export_File.replace('.xlsx' , '{}.xlsx'.format(td)))
            counter = time_Elapsed()
            today = datetime.datetime.today()   
            with open( os.path.join(log_D, log_Book_Name) , 'a', newline='' ) as myfile:   
                writer = csv.writer(myfile)
                writer.writerow([f'{today:%m_%d_%Y_%H_%M}', str(time_passed((time_Elapsed() 
                - PROGRAM_START_TIME))[0]) + "H", 'Backed up after {} min'.format(round(tst/60, 2))])
                myfile.close() 

        # Waits for specified time before sensing diode's current
        # while checking what IV to perform next --every second--.
        for x in range (0, math.ceil(abs(current_Sense_Frequency))):
            idle(0.985)
            if (arg ("states") ["run_Diode_Continuous"]) == "off": break

            dT_DEVICE_1 = dT_DEVICE_1 + 1; dT_DEVICE_2 = dT_DEVICE_2 + 1; dT_DEVICE_3 = dT_DEVICE_3 + 1
            
            """
            if 'on' in (list(arg("states").values())[1:]) :
            dT_DEVICE_1 = dT_DEVICE_1 + 1; dT_DEVICE_2 = dT_DEVICE_2 + 1; dT_DEVICE_3 = dT_DEVICE_3 + 1
            """
            
            check_What_IV_to_Run_Next(**arg("IV_to_Run_Next_Args"))

            if ((run_DEVICE_1_IV == True and (MD == "on" or MG =="on"))  or (run_DEVICE_2_IV == True 
            and (SD == "on" or SG =="on")) or (run_DEVICE_3_IV == True and AI == "on")): break
            


        if ((run_DEVICE_1_IV == True and (MD == "on" or MG =="on"))  or (run_DEVICE_2_IV == 
        True and (SD == "on" or SG =="on")) or (run_DEVICE_3_IV == True and AI == "on")): break
        
        if (arg ("states") ["run_Diode_Continuous"]) == "off": break 

    today = datetime.datetime.today()   
    with open( os.path.join(log_D, log_Book_Name) , 'a', newline='' ) as myfile:   
        writer = csv.writer(myfile)
        writer.writerow([f'{today:%m_%d_%Y_%H_%M}', str(time_passed((time_Elapsed() 
        - PROGRAM_START_TIME))[0]) + "H", f"Ending {measurement}",  str(int(time_Elapsed()- start_Time))])
        myfile.close()
    

    print("DEVICE_1 time passed " + str(dT_DEVICE_1))
    print("DEVICE_2 time passed " + str(dT_DEVICE_2))
    print("DEVICE_3 time passed " + str(dT_DEVICE_3))
  

while (True):
    
    if arg("ON-OFF")["state"] == "OFF" or arg("ON-OFF")["state"] == "off":
        mtrx.write('*RST')                                  # Resets the switching matrix
        mtrx.write('*RST')                                  # Resets the switching matrix
        break

    if (arg ("states") ["run_DEVICE_1_IV"]) == "on" and run_DEVICE_1_IV == True:

        ID_VD(**arg("DEVICE_1_IDVD_Args"))
        ID_VG(**arg("DEVICE_1_IDVG_Args"))
        check_What_IV_to_Run_Next(**arg("IV_to_Run_Next_Args"))

    if (arg ("states") ["run_DEVICE_2_IV"]) == "on" and run_DEVICE_2_IV == True:

        ID_VD (**arg("DEVICE_2_IDVD_Args"))
        ID_VG (**arg("DEVICE_2_IDVG_Args"))
        check_What_IV_to_Run_Next(**arg("IV_to_Run_Next_Args"))

    if (arg ("states") ["run_DEVICE_3_IV"]) == "on" and run_DEVICE_3_IV == True:
        print("f")
        diode_IVs (**arg("diode_IVs_Args"))
        check_What_IV_to_Run_Next(**arg("IV_to_Run_Next_Args"))

    if (arg ("states") ["run_Diode_Continuous"]) == "on":
        print("h")
        diode_Continuous(**arg("diode_Continuous_Args"))    

    elif (arg ("states") ["run_Diode_Continuous"]) == "off":
        
        smu.write('*rst')                                   # Resets the sourcemeter
        smu.write('opc?')                                   # Waits for all commands to be executed before executing next commands
        mtrx.write('*RST')                                  # Resets the switching matrix
        mtrx.write('opc?')                                  # Waits for all commands to be executed before executing next commands
        
        today = datetime.datetime.today()
        with open( os.path.join(log_D, log_Book_Name) , 'a', newline='' ) as myfile:   
            writer = csv.writer(myfile)
            writer.writerow([f'{today:%m_%d_%Y_%H_%M}', "Continuous diode voltage is off"])
            myfile.close()
            
        while True:
            idle(0.985)
            
            MD, MG, SD, SG, AI = (arg("DEVICE_1_IDVD_Args")["state"], arg("DEVICE_1_IDVG_Args")["state"],
            arg("DEVICE_2_IDVD_Args")["state"], arg("DEVICE_2_IDVG_Args")["state"],arg("diode_IVs_Args")["state"])

            dT_DEVICE_1 = dT_DEVICE_1 + 1; dT_DEVICE_2 = dT_DEVICE_2 + 1; dT_DEVICE_3 = dT_DEVICE_3 + 1   
                                           
            check_What_IV_to_Run_Next(**arg("IV_to_Run_Next_Args"))
            
            if ((run_DEVICE_1_IV == True and (MD == "on" or MG =="on"))  or (run_DEVICE_2_IV == True 
            and (SD == "on" or SG =="on")) or (run_DEVICE_3_IV == True and AI == "on")): break

            if (arg ("states") ["run_Diode_Continuous"]) == "on": break
           